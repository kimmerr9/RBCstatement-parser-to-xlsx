import re
from pdfminer.high_level import extract_pages
from pdfminer.layout import LTChar
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from datetime import datetime, date


def extract(filename: str):
    """
    Return a list where each index is all the text on one page.
    i.e. if you have 5 pages, list will have 5 items
    """
    pages = extract_pages(filename)
    page_content = []
    for figures in pages:
        page_text = ''
        for elems in figures:
            for c in elems:
                if isinstance(c, LTChar):
                    page_text += c.get_text()
        page_content.append(page_text)
    return page_content


def format(page_content):
    """
    For extract() content, will iterate through printing the statements grouping them in one list
    """
    pattern = re.compile(
        r'(?:\(\$\))?'
        r'([A-Z]{3}\s\d{2})'
        r'[A-Z]{3}\s\d{2}'
        r'(.*?)'    
        r'(?:EON|\s(?:ON|QC))\s*\d*'
        r'\$(\d+\.\d{2})'
    )
    major = []
    minor = []
    for i in range(len(page_content)):
        matches = pattern.findall(page_content[i])
        major.append(matches)
    
    for j in range(len(major)):
        for k in range(len(major[j])):
            minor.append(major[j][k])

    return minor

def to_excel(filename):
    content = extract(filename)
    transactions = format(content)
    wb = Workbook()
    ws = wb.active
    ws.title = "statement"

    headers = ["Date", "Transaction", "Amount"]
    ws.append(headers)

    for col in range(1,4):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")   


    current_year = date.today().year
    for date_str, name, amount_str in transactions:
        date_obj = datetime.strptime(f"{current_year} {date_str}", "%Y %b %d").date()
        amount_float = float(amount_str.replace("$", ""))
        ws.append([date_obj, name, amount_float])

    for row in ws.iter_rows(min_row=2, min_col=1, max_col=1):
        for cell in row:
            cell.number_format = 'yyyy-mm-dd'

    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 12

    wb.save("transactions.xlsx")

if __name__ == '__main__':
    # extracttt = extract("statement_mai.pdf")
    # formattt = format(extracttt)
    # print(formattt)

    to_excel("statement_mai.pdf")